import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 兼容 Windows 和 Linux/macOS 的文件路径输入
def get_valid_filepath(filename):
    """获取绝对路径，确保兼容不同操作系统"""
    filepath = os.path.abspath(filename)
    if not os.path.isfile(filepath):
        print(f"错误：文件 {filename} 不存在，请检查文件路径！")
        sys.exit(1)
    return filepath

# 用户输入 Excel 文件名（兼容不同系统）
file1 = input("请输入第一个 Excel 文件名（含扩展名）：").strip()
file2 = input("请输入第二个 Excel 文件名（含扩展名）：").strip()
# 输出文件名组合 file1_file2_deduplicated.xlsx
output_file = f"{file1.split('.')[0]}_{file2.split('.')[0]}_deduplicated.xlsx"
try:
    # 获取绝对路径，确保不同系统兼容
    file1 = get_valid_filepath(file1)
    file2 = get_valid_filepath(file2)
    output_file = os.path.abspath(output_file)

    # 读取 Excel 文件
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # 检查两个文件是否有相同的列
    common_columns = list(set(df1.columns) & set(df2.columns))
    if not common_columns:
        print("错误：两个 Excel 文件没有共同的列，无法匹配数据。")
        sys.exit(1)

    print(f"检测重复数据，基于公共列：{common_columns}")

    # 查找重复数据（基于公共列匹配）
    duplicates = df1.merge(df2, on=common_columns, how="inner")

    if duplicates.empty:
        print("未找到重复数据。")
    else:
        # 保存到临时文件
        temp_file = os.path.abspath("temp_output.xlsx")
        duplicates.to_excel(temp_file, index=False)

        # 读取 Excel 并应用样式（标红）
        wb = load_workbook(temp_file)
        ws = wb.active
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        # 遍历行并标记重复行
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.fill = red_fill

        # 保存最终文件
        wb.save(output_file)
        # 删除临时文件
        os.remove(temp_file)
        print(f"✅ 重复数据已标记并保存到 {output_file}")

except Exception as e:
    print(f"发生错误：{e}")
    sys.exit(1)

